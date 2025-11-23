"""
Servicio de generación de reportes PDF y Excel para LOGIFARMA PQR
"""
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
import matplotlib
matplotlib.use('Agg')  # Backend sin GUI
import matplotlib.pyplot as plt
import io
from datetime import datetime
from typing import List, Dict, Any

class ReportesService:
    """Servicio para generar reportes en PDF y Excel"""

    @staticmethod
    def generar_grafica_barras(datos: List[Dict], xlabel: str, ylabel: str, title: str) -> io.BytesIO:
        """Genera una gráfica de barras y la retorna como BytesIO"""
        plt.figure(figsize=(10, 6))
        labels = [d['label'] for d in datos]
        values = [d['value'] for d in datos]

        plt.bar(labels, values, color='#059669')
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.title(title)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()

        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        plt.close()

        return buffer

    @staticmethod
    def generar_grafica_linea(datos: List[Dict], xlabel: str, ylabel: str, title: str) -> io.BytesIO:
        """Genera una gráfica de línea y la retorna como BytesIO"""
        plt.figure(figsize=(12, 6))
        labels = [d['label'] for d in datos]
        values = [d['value'] for d in datos]

        plt.plot(labels, values, marker='o', color='#059669', linewidth=2, markersize=8)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.title(title)
        plt.xticks(rotation=45, ha='right')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()

        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        plt.close()

        return buffer

    @staticmethod
    def generar_pdf_desempeno_agentes(datos: Dict, fecha_inicio: str, fecha_fin: str) -> io.BytesIO:
        """Genera reporte PDF de desempeño de agentes"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#059669'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Título
        story.append(Paragraph("REPORTE DE DESEMPEÑO DE AGENTES", title_style))
        story.append(Paragraph(f"Período: {fecha_inicio} a {fecha_fin}", styles['Normal']))
        story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Tabla de datos
        if datos.get('agentes'):
            table_data = [['Agente', 'Abiertos', 'Cerrados', 'Promedio (hrs)']]
            for agente in datos['agentes']:
                table_data.append([
                    agente['agente'],
                    str(agente['abiertos']),
                    str(agente['cerrados']),
                    f"{agente['promedio_horas']:.2f}"
                ])

            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(t)

        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generar_excel_desempeno_agentes(datos: Dict, fecha_inicio: str, fecha_fin: str) -> io.BytesIO:
        """Genera reporte Excel de desempeño de agentes"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Desempeño Agentes"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezados
        ws['A1'] = "REPORTE DE DESEMPEÑO DE AGENTES"
        ws['A1'].font = Font(bold=True, size=16, color="059669")
        ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A3'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Tabla de datos
        headers = ['Agente', 'Casos Abiertos', 'Casos Cerrados', 'Promedio Horas']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Datos
        if datos.get('agentes'):
            for row_idx, agente in enumerate(datos['agentes'], start=6):
                ws.cell(row=row_idx, column=1, value=agente['agente']).border = border
                ws.cell(row=row_idx, column=2, value=agente['abiertos']).border = border
                ws.cell(row=row_idx, column=3, value=agente['cerrados']).border = border
                ws.cell(row=row_idx, column=4, value=round(agente['promedio_horas'], 2)).border = border

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generar_pdf_casos_periodo(datos: Dict, fecha_inicio: str, fecha_fin: str) -> io.BytesIO:
        """Genera reporte PDF de casos por período"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#059669'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Título
        story.append(Paragraph("REPORTE DE CASOS POR PERÍODO", title_style))
        story.append(Paragraph(f"Período: {fecha_inicio} a {fecha_fin}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Resumen ejecutivo
        story.append(Paragraph("<b>Resumen Ejecutivo</b>", styles['Heading2']))
        resumen_data = [
            ['Total de Casos', str(datos.get('total_casos', 0))],
            ['Casos Abiertos', str(datos.get('abiertos', 0))],
            ['Casos Cerrados', str(datos.get('cerrados', 0))],
            ['Casos en Proceso', str(datos.get('en_proceso', 0))],
            ['Tiempo Promedio Resolución', f"{datos.get('tiempo_promedio', 0):.2f} horas"]
        ]

        t = Table(resumen_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        story.append(t)

        doc.build(story)
        buffer.seek(0)
        return buffer
